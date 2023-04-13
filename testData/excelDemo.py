import openpyxl

book = openpyxl.load_workbook("D:\\Selenium-Python\\PytestAFramework\\excel\\pythonDemo.xlsx")
sheet = book.active
Dict = {}
# cell = sheet.cell(row=1, column=2)
# print(cell.value)

# sheet.cell(row=1, column=2).value = "Testing"

# print(sheet.cell(row=1, column=2).value)
#
# print(sheet.max_row)
#
# print(sheet.max_column)
#
# print(sheet["A5"].value)

for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "TC1":

        for j in range(2, sheet.max_column + 1):  # to get columns

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
